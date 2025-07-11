# Organizador de horários das disciplinas da UFRB para pessoas ansiosas
# (espera a matrícula começar, rapaz...)

import os
import re
import sys
import time
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# --- Configuração Inicial ---
if getattr(sys, 'frozen', False):  # Executável
    script_dir = os.path.dirname(sys.executable)
else:  # Script .py normal
    script_dir = os.path.dirname(os.path.abspath(__file__))

os.chdir(script_dir)

# Criar diretórios necessários
PASTA_SAIDA = os.path.join(script_dir, "Planilhas")
PASTA_DADOS = os.path.join(script_dir, "Dados")

os.makedirs(PASTA_SAIDA, exist_ok=True)
os.makedirs(PASTA_DADOS, exist_ok=True)

ARQUIVO_HORARIOS = os.path.join(PASTA_DADOS, "dados_horarios.csv")
ARQUIVO_DISCIPLINAS = os.path.join(PASTA_DADOS, "dados_disciplinas.csv")

# --- Constantes e Mapeamentos ---
dias_semana = {'2': 'SEGUNDA', '3': 'TERÇA', '4': 'QUARTA', '5': 'QUINTA', '6': 'SEXTA', '7': 'SÁBADO'}
ordem_dias = ['SEGUNDA', 'TERÇA', 'QUARTA', 'QUINTA', 'SEXTA', 'SÁBADO']
turnos = {'M': 'Manhã', 'T': 'Tarde', 'N': 'Noite'}
horarios_turno = {
    'M': {'1': '7 às 8', '2': '8 às 9', '3': '9 às 10', '4': '10 às 11', '5': '11 às 12'},
    'T': {'1': '13 às 14', '2': '14 às 15', '3': '15 às 16', '4': '16 às 17', '5': '17 às 18'},
    'N': {'1': '18:30 às 19:30', '2': '19:30 às 20:30', '3': '20:30 às 21:30', '4': '21:30 às 22:30'}
}

# --- Funções Utilitárias ---
def limpar_tela():
    os.system('cls' if os.name == 'nt' else 'clear')

def interpretar_codigo_bloco(bloco):
    for i, c in enumerate(bloco):
        if c in turnos:
            dias = [dias_semana[d] for d in bloco[:i] if d in dias_semana]
            turno = bloco[i]
            horarios = [horarios_turno[turno][h] for h in bloco[i+1:] if h in horarios_turno[turno]]
            return dias, horarios
    raise ValueError("Turno não identificado no bloco.")

def limpar_todos_os_dados():
    apagado = False
    for caminho in [ARQUIVO_HORARIOS, ARQUIVO_DISCIPLINAS]:
        if os.path.exists(caminho):
            os.remove(caminho)
            apagado = True
        else:
            apagado = False
    if apagado:
        print("\nTodos os dados foram apagados com sucesso!")
    else:
        print("\nNenhum dado existente para ser apagado.")
    
    input("\nPressione Enter para voltar ao menu principal...")
    return pd.DataFrame(columns=["Disciplina", "Dia", "Horário"]), pd.DataFrame(columns=["ID", "Nome"]), 1

def gerar_nome_proximo_arquivo(base_nome="TABELA DE HORÁRIOS", extensao=".xlsx"):
    padrao = re.compile(rf"{re.escape(base_nome)} \(_(\d+)_\){re.escape(extensao)}")
    arquivos_existentes = [f for f in os.listdir(PASTA_SAIDA) if padrao.match(f)]
    if not arquivos_existentes:
        return f"{base_nome} (_1_){extensao}"
    numeros = [int(padrao.match(f).group(1)) for f in arquivos_existentes]
    return f"{base_nome} (_{max(numeros)+1}_){extensao}"

def criar_grade_horaria(df_horarios_atual):
    # Estilos
    fonte_titulo = Font(name='Arial', size=12, bold=True, color="000000")
    fonte_turno = Font(name='Arial', size=10, bold=True, color="FFFFFF")
    fonte_conteudo = Font(name='Arial', size=8, color="000000")
    fonte_padrao = Font(name='Arial', size=8)
    alinhamento_centro = Alignment(horizontal='center', vertical='center', wrap_text=True)
    preenchimento_cinza_claro = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
    preenchimento_cinza_escuro = PatternFill(start_color="606060", end_color="606060", fill_type="solid")
    preenchimento_branco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    borda_fina = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Grade Horária"

    # Ajuste de dimensões
    for i in range(1, 19):
        for j in range(1, 8):
            cel = ws.cell(row=i, column=j)
            cel.font = fonte_padrao
        ws.row_dimensions[i].height = 25
    for i in [2, 8, 14]:
        ws.row_dimensions[i].height = 20
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 25

    # Cabeçalho dos dias
    ws.cell(row=1, column=1, value="HORÁRIOS")
    for col in range(1, 8):
        cel = ws.cell(row=1, column=col)
        cel.font = fonte_titulo
        cel.alignment = alinhamento_centro
        cel.fill = preenchimento_cinza_claro
        cel.border = borda_fina
    for j, dia in enumerate(ordem_dias):
        ws.cell(row=1, column=j + 2).value = dia.upper()

    # Estrutura por turno
    turnos_linhas = {
        "MANHÃ": (2, range(3, 8), ["7 às 8", "8 às 9", "9 às 10", "10 às 11", "11 às 12"]),
        "TARDE": (8, range(9, 14), ["13 às 14", "14 às 15", "15 às 16", "16 às 17", "17 às 18"]),
        "NOITE": (14, range(15, 19), ["18:30 às 19:30", "19:30 às 20:30", "20:30 às 21:30", "21:30 às 22:30"])
    }

    for turno, (linha_turno, linhas_horarios, horarios) in turnos_linhas.items():
        for col in range(1, 8):
            cel = ws.cell(row=linha_turno, column=col)
            cel.fill = preenchimento_cinza_escuro
            cel.border = borda_fina
            cel.alignment = alinhamento_centro
        ws.cell(row=linha_turno, column=1, value=turno).font = fonte_turno

        for idx, linha in enumerate(linhas_horarios):
            hora = horarios[idx]
            ws.cell(row=linha, column=1, value=hora).font = fonte_titulo
            ws.cell(row=linha, column=1).alignment = alinhamento_centro
            ws.cell(row=linha, column=1).fill = preenchimento_cinza_claro
            ws.cell(row=linha, column=1).border = borda_fina

            for col, dia in enumerate(ordem_dias, start=2):
                cel = ws.cell(row=linha, column=col)
                cel.fill = preenchimento_branco
                cel.alignment = alinhamento_centro
                cel.border = borda_fina
                entrada = df_horarios_atual[
                    (df_horarios_atual['Dia'] == dia) & (df_horarios_atual['Horário'] == hora)
                ]
                if not entrada.empty:
                    cel.value = entrada.iloc[0]['Disciplina']
                    cel.font = fonte_conteudo

    caminho = os.path.join(PASTA_SAIDA, gerar_nome_proximo_arquivo())
    wb.save(caminho)
    print(f"\nPlanilha gerada: '{os.path.basename(caminho)}'")
    print(f"Localização: {os.path.abspath(caminho)}\n")

def gerar_nome_proximo_arquivo(base_nome="PLANILHA DE HORÁRIOS", extensao=".xlsx"):
    padrao = re.compile(rf"{re.escape(base_nome)} \(_(\d+)_\){re.escape(extensao)}")
    arquivos_existentes = [f for f in os.listdir(PASTA_SAIDA) if padrao.match(f)]
    if not arquivos_existentes:
        return f"{base_nome} (_1_){extensao}"
    numeros = [int(padrao.match(f).group(1)) for f in arquivos_existentes]
    proximo_num = max(numeros) + 1
    return f"{base_nome} (_{proximo_num}_){extensao}"

# Funções de Menu

def exibir_disciplinas_cadastradas(df_disciplinas):
    print("\n" + "=" * 40)
    print("Disciplinas Cadastradas:")
    if df_disciplinas.empty:
        print("Nenhuma disciplina cadastrada.")
    else:
        print(df_disciplinas.to_string(index=False))
    print("=" * 40)

def menu_cadastrar(df_horarios, df_disciplinas, proximo_id):
    while True:
        disciplina_nome = input("\nDigite o nome da nova disciplina (ou 'sair' para retornar): ").strip()
        if disciplina_nome.lower() == 'sair':
            break
        if disciplina_nome in df_disciplinas['Nome'].values:
            print(f"Erro: A disciplina '{disciplina_nome}' já está cadastrada.")
            continue
        codigos = input(f"Cole ou digite o(s) código(s) de horário para a disciplina '{disciplina_nome}' (ex: 24T12 6M345): ").upper().strip()
        blocos, conflito, novas_linhas = codigos.split(), False, []
        for bloco in blocos:
            try:
                dias, horarios = interpretar_codigo_bloco(bloco)
                for dia in dias:
                    for horario in horarios:
                        existe = df_horarios[(df_horarios['Dia'] == dia) & (df_horarios['Horário'] == horario)]
                        if not existe.empty:
                            print(f"Conflito com '{existe.iloc[0]['Disciplina']}' em {dia} - {horario}")
                            conflito = True
                        else:
                            novas_linhas.append({"Disciplina": disciplina_nome, "Dia": dia, "Horário": horario})
            except Exception as e:
                print(f"Erro no bloco '{bloco}': {e}"); conflito = True; break
        if not conflito and novas_linhas:
            df_horarios = pd.concat([df_horarios, pd.DataFrame(novas_linhas)], ignore_index=True)
            df_disciplinas = pd.concat([df_disciplinas, pd.DataFrame([{'ID': proximo_id, 'Nome': disciplina_nome}])], ignore_index=True)
            proximo_id += 1
            print(f"Disciplina '{disciplina_nome}' cadastrada com sucesso!")
        else:
            print("Cadastro cancelado por conflitos ou erros.")
    return df_horarios, df_disciplinas, proximo_id

def menu_excluir(df_horarios, df_disciplinas):
    while True:
        exibir_disciplinas_cadastradas(df_disciplinas)
        if df_disciplinas.empty:
            input("\nPressione Enter para voltar ao menu principal...")
            break
        id_str = input("Digite o número de ID da disciplina a ser excluída (ou 'sair' para retornar): ").strip()
        if id_str.lower() == 'sair':
            break
        if id_str.isdigit():
            id_para_apagar = int(id_str)
            disciplina_row = df_disciplinas[df_disciplinas['ID'] == id_para_apagar]
            if not disciplina_row.empty:
                nome_disciplina = disciplina_row.iloc[0]['Nome']
                df_disciplinas = df_disciplinas[df_disciplinas['ID'] != id_para_apagar]
                df_horarios = df_horarios[df_horarios['Disciplina'] != nome_disciplina]
                print(f"\nDisciplina '{nome_disciplina}' (ID: {id_para_apagar}) foi removida.")
            else:
                print(f"\nErro: ID '{id_para_apagar}' não encontrado.")
        else:
            print("\nEntrada inválida. Por favor, digite um número de ID válido.")
    return df_horarios, df_disciplinas

def main():
    df_horarios = pd.read_csv(ARQUIVO_HORARIOS) if os.path.exists(ARQUIVO_HORARIOS) else pd.DataFrame(columns=["Disciplina", "Dia", "Horário"])
    df_disciplinas = pd.read_csv(ARQUIVO_DISCIPLINAS) if os.path.exists(ARQUIVO_DISCIPLINAS) else pd.DataFrame(columns=["ID", "Nome"])
    proximo_id = df_disciplinas['ID'].max() + 1 if not df_disciplinas.empty else 1

    try:
        while True:
            limpar_tela()
            print("╔═════════════════════════════════════╗")
            print("║           MENU PRINCIPAL            ║")
            print("╠═════════════════════════════════════╣")
            print("║ 1. Cadastrar Disciplina             ║")
            print("║ 2. Disciplinas Cadastradas          ║")
            print("║ 3. Excluir Disciplina               ║")
            print("║ 4. Limpar Dados                     ║")
            print("║ 5. Sair e Gerar Planilha            ║")
            print("╚═════════════════════════════════════╝")
            print("Feito por: @dj_britto0")
            opcao = input("\nEscolha uma opção: ").strip()
            if opcao == '1':
                df_horarios, df_disciplinas, proximo_id = menu_cadastrar(df_horarios, df_disciplinas, proximo_id)
            elif opcao == '2':
                exibir_disciplinas_cadastradas(df_disciplinas)
                input("\nPressione Enter para voltar ao menu principal...")
            elif opcao == '3':
                df_horarios, df_disciplinas = menu_excluir(df_horarios, df_disciplinas)
            elif opcao == '4':
                confirmacao = input("\nATENÇÃO! Isso apagará todos os dados. Tem certeza? (sim/nao): ").lower()
                while confirmacao not in ['sim', 's', 'nao', 'n']:
                    confirmacao = input("Por favor, digite 'sim' ou 'nao': ").lower()
                if confirmacao in ['sim', 's']:
                    df_horarios, df_disciplinas, proximo_id = limpar_todos_os_dados()
                else:
                    print("\nOperação cancelada.")
                    input("\nPressione Enter para voltar ao menu principal...")
            elif opcao == '5':
                print("\nEncerrando e salvando alterações...")
                time.sleep(2)
                break
            else:
                print("\nOpção inválida! Por favor, escolha um número do menu.")
    finally:
        if not df_horarios.empty:
            df_horarios.to_csv(ARQUIVO_HORARIOS, index=False)
            df_disciplinas.to_csv(ARQUIVO_DISCIPLINAS, index=False)
            criar_grade_horaria(df_horarios)
            time.sleep(5)
        else:
            print("\nNão há dados cadastrados. Nenhum arquivo foi salvo.")
            time.sleep(3)

if __name__ == "__main__":
    main()
